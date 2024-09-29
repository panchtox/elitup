import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from datetime import datetime

def generate_report(articles, evidence):
    wb = Workbook()
    
    # Evidence sheet
    ws_evidence = wb.active
    ws_evidence.title = "Evidence"
    headers = ["Owner", "País", "Producto", "Search Strategy", "Scope", "Search URL", "Search Date", "Articles Number"]
    ws_evidence.append(headers)
    for e in evidence:
        ws_evidence.append([e.owner, e.pais, e.producto, e.searchStrategy, e.scope, e.searchUrl, e.searchDate, e.articles_number])

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    total_articles = len(articles)
    relevant_articles = len([a for a in articles if a.status == "Relevante"])
    reportable_articles = len([a for a in articles if a.status == "Reportable"])
    no_relevante_articles = len([a for a in articles if a.status == "No relevante"])
    
    ws_summary.append(["Report Summary"])
    ws_summary.append(["Generated on", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["Total Articles", total_articles])
    ws_summary.append(["Relevant Articles", relevant_articles])
    ws_summary.append(["Reportable Articles", reportable_articles])
    ws_summary.append(["No Relevante Articles", no_relevante_articles])
    
    # Add chart for article distribution
    pie = PieChart()
    labels = Reference(ws_summary, min_col=1, min_row=4, max_row=6)
    data = Reference(ws_summary, min_col=2, min_row=3, max_row=6)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Article Distribution"
    ws_summary.add_chart(pie, "A9")

    # Detail sheet
    ws_detail = wb.create_sheet("Detail")
    headers = ["Título", "Abstract", "Fecha de hit", "sourceUrl", "Categoría"]
    ws_detail.append(headers)
    
    for a in articles:
        if a.status in ["Relevante", "Reportable"]:
            abstract = a.spanishAbstract or a.englishAbstract or a.portugueseAbstract or ""
            ws_detail.append([a.title, abstract, a.dateOfHit, a.sourceUrl, a.status])

    # No Relevante sheet
    ws_no_relevante = wb.create_sheet("No Relevante")
    ws_no_relevante.append(headers)
    
    for a in articles:
        if a.status == "No relevante":
            abstract = a.spanishAbstract or a.englishAbstract or a.portugueseAbstract or ""
            ws_no_relevante.append([a.title, abstract, a.dateOfHit, a.sourceUrl, a.status])

    # Style the sheets
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    # Adjust row height for wrapped text
    for ws in [ws_evidence, ws_detail, ws_no_relevante]:
        for row in ws.iter_rows(min_row=2):
            max_height = 0
            for cell in row:
                if cell.value:
                    text_lines = str(cell.value).count('\n') + 1
                    max_height = max(max_height, 15 * text_lines)
            ws.row_dimensions[row[0].row].height = max_height

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output
